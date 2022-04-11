import schedule
import time
import datetime as dt
from pymongo import MongoClient
import datetime as dt
from openpyxl import Workbook
import win32com.client

start = dt.datetime.now()


def shipping_report():
    current_date = (dt.date.today())
    #start = dt.datetime.combine((current_date) - dt.timedelta(days=1), dt.time(16,30,0,0))
    #end = dt.datetime.combine(current_date, dt.time(16,30,0,0))


    start = dt.datetime(2021, 6, 22, 0, 0, 0, 0)
    end = dt.datetime(2021, 6, 22, 23, 59, 59, 999999)

    client = MongoClient('localhost', 27017)
    db = client.signode
    
    
    string_array = [db.markham, db.glenview, db.surrey]
    location_array = ['markham','glenview', 'surrey']
    check_array = [1,2,3]

    workbook = Workbook()
    sheet = workbook.active
    p = 0
    
    for location in string_array:
        no = (location.count_documents({'shippedDate': {'$gte': start, '$lt': end}}))
        x = 0
        check = True
        print(no)
        Ship_via = [0] * no
        POs = [0] * no
        Order = [0] * no
        Ship_to = [0] * no
        Ship_date = [0] * no
        Recieve_date = [0] * no
        for collection in location.find({'shippedDate': {'$gte': start, '$lt': end}}):
            Order[x] = collection['_id']
            POs[x] = collection['PO']
            Ship_via[x] = collection['via']
            Ship_to[x] = collection['shipTo']
            Ship_date[x] = (str(collection['shippedDate']))[:16]
            Recieve_date[x] = (str(collection['dateReceived']))[:16]

            x = x + 1

        data = []
        for i in range(no):
                data.append(
                    {
                        "Order #" : Order[i],
                        "PO #": POs[i],
                        "Ship To" : Ship_to[i],
                        "Ship Via" : Ship_via[i],
                        "Date Received" : Recieve_date[i],
                        "Ship Date" : Ship_date[i]
                    }
                )
        
        sheet.title = location_array[p]
        

        sheet["A1"] = "Order #"
        sheet["B1"] = "PO #"
        sheet["C1"] = "Ship TO"
        sheet["D1"] = "Ship Via"
        sheet["E1"] = "Date Received"
        sheet["F1"] = "Date Shipped"
        

        MIN_ROW = 2
        MAX_ROW = len(data) + 1
        MIN_COL = 1
        MAX_COL = 6


        for item, row in enumerate(sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=False)):
                        row[1].value = data[item]['PO #']
                        row[2].value = data[item]['Ship To']
                        row[3].value = data[item]['Ship Via']
                        row[4].value = data[item]['Date Received']
                        row[5].value = data[item]['Ship Date']
                        row[0].value = data[item]['Order #']
                        row[0].hyperlink = "http://10.100.6.111:3000/orders/" + data[item]['Order #']
                        row[0].style = "Hyperlink"

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
        
        if len(data) == 0:
            check = False

        check_array[p] = check
        p = p+1
        if p != 3:
            sheet = workbook.create_sheet()



    name = str(current_date) + " - Shipping Report.xlsx"
    workbook.save(filename=name)

    juice = False

    for x in check_array:
        if x:
            juice = True



    if juice:
        outlook = win32com.client.Dispatch('outlook.application')

        mail = outlook.CreateItem(0)

        mail.To = '*Email*'
        mail.Subject = 'Shipping Report for: ' + str(current_date)
        mail.Body = 'Please find the list of Shipped Orders for '+ str(current_date) + ' attached above. '
        mail.Attachments.Add('*File Path*' + name)
        mail.Send()
    


schedule.every().day.at("16:44").do(shipping_report)



while True:
    schedule.run_pending()
    time.sleep(1)
